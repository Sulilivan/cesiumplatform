import pandas as pd
from sqlalchemy.orm import Session
from sql_app.database import SessionLocal
from sql_app.models import MonitorPoint, Measurement
from datetime import datetime
import openpyxl

def parse_position(position_str):
    """
    解析平面位置字符串，提取相对坐标
    """
    if pd.isna(position_str):
        return 0, 0
    
    BASE_LAT = 30.0
    BASE_LON = 120.0
    
    try:
        if '坝左' in position_str:
            parts = position_str.replace('坝左', '').split('+')
            x_offset = -float(parts[1]) if len(parts) > 1 else 0
            y_offset = 0
        elif '坝右' in position_str:
            parts = position_str.replace('坝右', '').split('+')
            x_offset = float(parts[1]) if len(parts) > 1 else 0
            y_offset = 0
        else:
            x_offset, y_offset = 0, 0
        
        lat = BASE_LAT + (y_offset / 111000)
        lon = BASE_LON + (x_offset / (111000 * 0.77))
        
        return lon, lat
    except:
        return BASE_LON, BASE_LAT

def import_inverted_plumb_data():
    """从监测资料.xlsx的倒垂线工作表导入数据"""
    print("="*60)
    print("开始导入倒垂线数据")
    print("="*60 + "\n")
    
    file_path = r'D:\desktop\新建文件夹 (2)\SmartWaterPlatform\backend\data\监测资料.xlsx'
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb['倒垂线']
        
        db = SessionLocal()
        
        try:
            # 获取测点信息（从第5行）
            # 列映射：IP1在第4列，IP3在第6列，IP5在第8列，IP9在第10列，IP7在第12列，IP8在第14列，IP6在第16列，IP4在第18列，IP2在第20列
            point_columns = {
                4: ('IP1', '坝上0+000.900 坝左0+199.600', 153),
                6: ('IP3', '坝下0+007.600 坝左0+161.800', 120),
                8: ('IP5', '坝下0+011.200 坝左0+109.600', 88),
                10: ('IP9', '坝下0+004.700   坝左0+046.000', 43),
                12: ('IP7', '坝下0+003.500坝右0+008.300', 43),
                14: ('IP8', '坝下0+002.000坝右0+092.000', 60),
                16: ('IP6', '坝下0+012.500 坝右0+177.650', 88),
                18: ('IP4', '坝下0+006.500 坝右0+257.000', 120),
                20: ('IP2', '坝上0+001.500 坝右0+323.000', 153),
            }
            
            # 导入测点
            print("导入测点...")
            for col, (point_code, position, height) in point_columns.items():
                longitude, latitude = parse_position(position)
                
                existing_point = db.query(MonitorPoint).filter(
                    MonitorPoint.point_code == point_code
                ).first()
                
                if not existing_point:
                    point = MonitorPoint(
                        point_code=point_code,
                        point_name=point_code,
                        device_type='倒垂线',
                        longitude=longitude,
                        latitude=latitude,
                        height=height
                    )
                    db.add(point)
                    print(f"添加测点: {point_code}")
                else:
                    print(f"测点已存在: {point_code}")
            
            db.commit()
            print("测点导入完成\n")
            
            # 导入监测数据（从第7行开始）
            print("导入监测数据...")
            
            # 观测时间在第2列
            # 左右岸数据在偶数列，上下游数据在奇数列
            data_columns = {
                'IP1': (4, 5),
                'IP3': (6, 7),
                'IP5': (8, 9),
                'IP9': (10, 11),
                'IP7': (12, 13),
                'IP8': (14, 15),
                'IP6': (16, 17),
                'IP4': (18, 19),
                'IP2': (20, 21),
            }
            
            imported_count = 0
            for row_idx in range(7, ws.max_row + 1):
                # 获取观测时间（第2列）
                time_value = ws.cell(row=row_idx, column=2).value
                
                if time_value is None:
                    continue
                
                try:
                    if isinstance(time_value, str):
                        time = datetime.strptime(time_value, '%Y-%m-%d %H:%M:%S')
                    else:
                        time = time_value
                except:
                    continue
                
                # 为每个测点导入数据
                for point_code, (left_right_col, up_down_col) in data_columns.items():
                    # 左右岸数据
                    lr_value = ws.cell(row=row_idx, column=left_right_col).value
                    if lr_value is not None:
                        # 跳过公式和字符串
                        if isinstance(lr_value, str) and lr_value.startswith('='):
                            continue
                        try:
                            measurement = Measurement(
                                point_code=point_code,
                                value=float(lr_value),
                                time=time
                            )
                            db.add(measurement)
                            imported_count += 1
                        except (ValueError, TypeError):
                            pass
                    
                    # 上下游数据
                    ud_value = ws.cell(row=row_idx, column=up_down_col).value
                    if ud_value is not None:
                        # 跳过公式和字符串
                        if isinstance(ud_value, str) and ud_value.startswith('='):
                            continue
                        try:
                            measurement = Measurement(
                                point_code=point_code,
                                value=float(ud_value),
                                time=time
                            )
                            db.add(measurement)
                            imported_count += 1
                        except (ValueError, TypeError):
                            pass
                
                if imported_count % 1000 == 0:
                    print(f"已导入 {imported_count} 条数据...")
                    db.commit()
            
            db.commit()
            print(f"监测数据导入完成！共导入 {imported_count} 条数据\n")
            
        except Exception as e:
            db.rollback()
            print(f"导入数据时出错: {e}")
            import traceback
            traceback.print_exc()
        finally:
            db.close()
            
    except Exception as e:
        print(f"读取文件失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import_inverted_plumb_data()
